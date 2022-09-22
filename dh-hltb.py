#!/usr/bin/env python3
'''
CLI tool for gathering howlongtobeat.com statistics for your games collection
tracked at databaze-her.cz.

Language note: All user-facing strings are written in Czech, because this script
has little use to anyone who doesn't understand Czech language.

Author: Kamil Páral <https://github.com/kparal/dh-hltb>
License: GNU AGPLv3+, see LICENSE file
'''
import sys
import argparse
import datetime
import pathlib
import os
import copy
import dataclasses
import csv
import collections
from typing import List, Iterable, Generator
import enum
import time
# external modules
import yaml
from bs4 import BeautifulSoup
from howlongtobeatpy import HowLongToBeat, HowLongToBeatEntry
import colorama
import pyexcel
import ezodf
import openpyxl

mapping_filename = 'mapping.yaml'
cache_filename = 'cache.yaml'

def parse_args():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description='''Zpracuje HTML stránky z databaze-her.cz (DH) a pro \
každou hru označenou jako "Chci si zahrát" nebo "Dohráno" zjistí doby hraní z \
howlongtobeat.com (HLTB). Výsledek uloží do XLSX/ODS/CSV souboru. Příklady URL \
k uložení jako vstupní HTML (nahraďte JMENO za vaše uživatelské jméno):

https://www.databaze-her.cz/uzivatele/JMENO/hry/chci-si-zahrat/?&razeni=3&styl=seznam&stranka=vse
https://www.databaze-her.cz/uzivatele/JMENO/hry/dohrane/?razeni=6&styl=seznam&stranka=vse

Se dvěma uloženými soubory pak lze spustit skript takto:
$ ./dh-hltb.py chci.html dohrane.html

Pro hry, které se na DH jmenují jinak než na HLTB, je nutné vytvořit záznam v \
souboru 'mapping.yaml'. Zkopírujte 'example-mapping.yaml' jako 'mapping.yaml' \
a upravte tak, aby obsahoval hry, na které se dotazujete (případně chcete \
ignorovat).
''')
    parser.add_argument('html_filenames', metavar='DH_HTML', nargs='+',
        help='uložené HTML soubory ze záložky "Hry" z databaze-her.cz. Více '\
            'souborů je zpracováno postupně, a vstupní řazení her je zachováno '\
            'i ve výstupním souboru.')
    parser.add_argument('-o', '--output', metavar='STATS', default='stats.xlsx',
        help='výstupní soubor ve formátu XLSX/ODS/CSV. Typ souboru se pozná '\
            'podle přípony. Nejlepší formátování má XLSX soubor. '\
            '[výchozí: %(default)s]')
    parser.add_argument('--cache-ttl', metavar='DAYS', type=int, default=30,
        help='údaje her starší než tento počet dní se načtou z HLTB znovu. '\
            'Hodnota 0 vynutí nové načtení vždy. [výchozí: %(default)s]')
    parser.add_argument('-i', '--include-ignored', action='store_true',
        default=False,
        help='zpracovat i hry označené jako ignorované v souboru '\
            f'{mapping_filename}')
    cachedir = os.path.join(pathlib.Path.home(), '.cache', 'dh-hltb')
    parser.add_argument('--cachedir', default=cachedir,
        help='adresář pro ukládání cache z HLTB [výchozí: %(default)s]')
    args = parser.parse_args()

    # validate
    if not (args.output.endswith('.xlsx') or args.output.endswith('.ods') or
            args.output.endswith('.csv')):
        parser.error('Podporované přípony výstupního souboru: ".xlsx" ".ods" '
            f'".csv". Zadáno: {args.output}')

    return args


@dataclasses.dataclass
class Game:
    dh_id: int = None
    title: str = None
    year: int = None
    wantplay: bool = None
    finished: bool = None
    finished_ts: str = None
    owned: bool = None

    hltb_id: int = None
    # time unit is implied to be 'hours'
    time_main: float = None
    time_extra: float = None
    time_complete: float = None
    time_all: float = None
    hltb_query_ts: str = None

    _cache_keys = [
        'hltb_id',
        'time_main',
        'time_extra',
        'time_complete',
        'time_all',
        'hltb_query_ts',
    ]


def parse_dh(html_filename: str) -> List[Game]:
    print(f'Procházím {html_filename} ...')
    with open(html_filename) as html_file:
        html = html_file.read()
    soup = BeautifulSoup(html, 'html.parser')

    # find index of 'Datum dohrání' column, if available
    dohrani_column = None
    headers = soup.select('div#user-games > div > table > thead > tr > th')
    dohrani = [h for h in headers
        if 'Datum' in h.stripped_strings and 'dohrání' in h.stripped_strings]
    if dohrani:
        assert len(dohrani) == 1
        dohrani = dohrani[0]
        dohrani_column = dohrani.parent.contents.index(dohrani)

    # parse all the games
    games = []
    dh_games = soup.select_one('div#user-games > div > table > tbody')
    for dh_game in dh_games.children:
        game = Game()
        game.dh_id = int(dh_game['data-id'])
        title_tag = dh_game.select_one('a.item-title')
        game.title = str(title_tag.string)
        year = title_tag.next_sibling.string.strip().replace('(','').replace(')','')
        if year.isdecimal():
            game.year = int(year)
        own_tag = dh_game.select_one('span.-own')
        game.owned = '-inactive' not in own_tag['class']
        wish_tag = dh_game.select_one('span.-wish')
        game.wantplay = '-inactive' not in wish_tag['class']
        finish_tag = dh_game.select_one('span.-finished')
        game.finished = '-inactive' not in finish_tag['class']
        if dohrani_column:
            dohrani = str(dh_game.contents[dohrani_column].string)
            game.finished_ts = datetime.datetime.strptime(
                dohrani, '%d.%m.%Y').strftime('%Y-%m-%d')

        if not game.wantplay and not game.finished:
            continue
        games.append(game)

    print(f'Nalezeno {len(games)} her')
    return games

def merge_game(entry1: Game, entry2: Game) -> Game:
    '''If there are two entries of the same game (same dh_id), try to merge
    their values and return the union of their data.'''
    if entry1.dh_id != entry2.dh_id:
        raise ValueError(f'Pokus o sjednocení dvou různých her:\n{entry1}\n'
            f'a\n{entry2}')
    entry = copy.deepcopy(entry1)
    for attr, val in entry1.__dict__.items():
        if val is None:
            setattr(entry, attr, getattr(entry2, attr))
    return entry

def create_dh_gamelist(html_filenames: List[str]) -> List[Game]:
    gamesdict = collections.OrderedDict()
    for html_filename in args.html_filenames:
        games = parse_dh(html_filename)
        for game in games:
            if game.dh_id in gamesdict:
                merged_game = merge_game(entry1=gamesdict[game.dh_id], entry2=game)
                gamesdict[game.dh_id] = merged_game
            else:
                gamesdict[game.dh_id] = game
    colorprint(Color.INFO, f'Celkem sesbíráno {len(gamesdict)} her')
    return gamesdict.values()


class HLTB():
    mapping_filename: str = None
    cache_filename: str = None
    args = None
    games: List[Game] = None
    cache: dict = {}
    mapping: dict = {}
    ignored: List[str] = []
    error_sleep_delay: float = 1

    def __init__(self, args, games: List[Game]):
        self.args = args
        self.games = games
        self.mapping_filename = os.path.join(find_prog_dir(), mapping_filename)
        self.cache_filename = os.path.join(self.args.cachedir, cache_filename)

    def query_hltb(self, game: Game, progress: str = '') -> (HowLongToBeatEntry):
        if game.dh_id in self.mapping:
            print(f'{progress} Zjišťuji: {game.title} (použito vlastní mapování) ...')
            hltb_id = self.mapping[game.dh_id]['hltb_id']
            title = self.mapping[game.dh_id]['hltb_title']
        else:
            print(f'{progress} Zjišťuji: {game.title} ...')
            hltb_id = None
            title = game.title

        tried_our_best = False
        while True:
            results = HowLongToBeat(input_minimum_similarity=0).search(
                game_name=title,
                similarity_case_sensitive=False)

            if results or tried_our_best:
                break

            if results is None:
                print_error('CHYBA SPOJENÍ NEBO NEPLATNÝ DOTAZ!')
                raise HLTBError

            if not results:
                title = self.more_searchable_name(title)
                colorprint(msg=f'Nic nenalezeno, zkouším hledat: {title}',
                           color=Color.INFO)
                tried_our_best = True
                continue

        if game.dh_id in self.ignored:
            # this happens if you force-include ignored games for querying
            print_args = {'color': Color.INFO}
            prefix = 'IGNOROVÁNO: '
        else:
            print_args = {'color': Color.ERROR, 'file': sys.stderr}
            prefix = ''

        if not results:
            colorprint(msg=f'{prefix}ŽÁDNÝ NÁLEZ PRO "{title}"! (DH ID: {game.dh_id})',
                       **print_args)
            time.sleep(self.error_sleep_delay)
            return None

        results = sorted(results, key=lambda x: x.similarity, reverse=True)

        if hltb_id:
            matches = [result for result in results if result.game_id == hltb_id]
            if len(matches) != 1:
                colorprint(msg=f'OČEKÁVÁN PŘESNĚ JEDEN NÁLEZ S HLTB ID {hltb_id}, ZÍSKÁNO '
                               f'{len(matches)}\nVÝSLEDKY HLEDÁNÍ "{title}":', **print_args)
                for result in results:
                    print('----')
                    print(self.format_result(result))
                print('----')
                time.sleep(self.error_sleep_delay)
                return None
            else:
                return matches[0]

        good_match_possible = True
        if len(results) >= 2 and results[0].similarity == results[1].similarity:
            # we can't reliably distinguish several matches with the same similarity
            good_match_possible = False
        if (results[0].game_name != game.title and
                results[0].game_alias != game.title and
                not self.equal_names([results[0].game_name, results[0].game_alias], [game.title])):
            # if the most similar result doesn't have the exact or almost equal
            # name, we can't say whether it is what we're looking for
            good_match_possible = False

        if not good_match_possible:
            colorprint(msg=f'{prefix}ŽÁDNÝ PŘESNÝ NÁLEZ PRO "{title}"! (DH ID: {game.dh_id})',
                       **print_args)
            for result in results:
                print('----')
                print(self.format_result(result))
            print('----')
            time.sleep(self.error_sleep_delay)
            return None

        # now we have a good match, results[0]
        best_match = results[0]

        # if we matched by title (not by id) and the title match was not exact,
        # inform the user
        if (best_match.game_name != game.title and
                best_match.game_alias != game.title and
                game.dh_id not in self.mapping):
            alias = f'(alias: {best_match.game_alias}) ' if best_match.game_alias else ''
            colorprint(msg=f'Přiřazen velice podobný název: {best_match.game_name} {alias}'
                           f'(rok vydání: {best_match.release_world})',
                       color=Color.INFO)

        return best_match

    @staticmethod
    def format_result(result: dict):
        gamesdict = collections.OrderedDict([
            ('Název', result.game_name),
            ('Alias', result.game_alias),
            ('Rok', result.release_world),
            ('Vývojář', result.profile_dev),
            ('Typ', result.game_type),
            ('HLTB ID', result.game_id),
            ('Podobnost', result.similarity),
        ])
        for key, value in gamesdict.copy().items():
            if not value:
                del gamesdict[key]

        return '\n'.join([f'{key}: {value}' for key, value in gamesdict.items()])

    def process_hltb_result(self, game: Game, hltb_result: HowLongToBeatEntry):
        game.time_main = hltb_result.main_story
        game.time_extra = hltb_result.main_extra
        game.time_complete = hltb_result.completionist
        game.time_all = hltb_result.all_styles
        game.hltb_id = hltb_result.game_id
        game.hltb_query_ts = datetime.datetime.now(
            datetime.timezone.utc).isoformat(timespec='seconds')

    def run(self):
        self.load_cache()
        self.load_mapping()
        if self.args.cache_ttl <= 0:
            print('Vynuceno obnovení informací pro všechny tituly')

        query_errors = 0
        for index, game in enumerate(self.games, start=1):
            progress_str = '({:={}}/{})'.format(index, len(str(len(self.games))),
                len(self.games))
            if game.dh_id in self.ignored and not self.args.include_ignored:
                print(f'{progress_str} Ignorováno: {game.title}')
                continue
            if not self.needs_refresh(game.hltb_query_ts):
                print(f'{progress_str} Aktuální: {game.title}')
                continue
            hltb_result = self.query_hltb(game, progress=progress_str)
            if hltb_result:
                self.process_hltb_result(game, hltb_result)
            if not hltb_result and game.dh_id not in self.ignored:
                query_errors += 1
        print(f'Zpracováno {len(self.games)} her')
        if query_errors:
            print_error(f'Počet chyb během zpracování: {query_errors}')

        self.save_cache()

    def save_cache(self):
        print('Ukládám dočasnou paměť ...')
        # update cache with fresh game info
        for game in self.games:
            # skip completely empty entries, no reason to cache empty data
            if not any([getattr(game, key) for key in Game._cache_keys]):
                continue
            cacheitem = {}
            for key in Game._cache_keys:
                cacheitem[key] = getattr(game, key)
            # just to make the cache easily inspectable
            cacheitem['dh_title'] = game.title

            self.cache[game.dh_id] = cacheitem
        # save cache
        os.makedirs(self.args.cachedir, exist_ok=True)
        with open(self.cache_filename, mode='w') as cache_file:
            cache_file.write(
                yaml.safe_dump(self.cache)
            )
        print(f'Dočasná paměť uložena do {self.cache_filename}')

    def load_cache(self):
        print(f'Nahrávám dočasnou paměť z {self.cache_filename} ...')
        if not os.path.exists(self.cache_filename):
            print('Soubor dočasné paměti nenalezen, přeskakuji')
            return
        with open(self.cache_filename) as cache_file:
            self.cache = yaml.safe_load(
                cache_file.read()
            )
        # drop outdated cache items, to avoid growing cache indefinitely
        for dh_id, cacheitem in self.cache.copy().items():
            if self.needs_refresh(cacheitem['hltb_query_ts']):
                del self.cache[dh_id]
        # update current game list
        for game in self.games:
            cachevals = self.cache.get(game.dh_id)
            if not cachevals:
                continue
            for key in Game._cache_keys:
                if key in cachevals:
                    setattr(game, key, cachevals[key])
        print('Dočasná paměť nahrána')

    def load_mapping(self):
        print(f'Nahrávám vlastní mapování ID her z {self.mapping_filename} ...')
        if not os.path.exists(self.mapping_filename):
            print('Mapovací soubor nenalezen, přeskakuji')
            return
        with open(self.mapping_filename) as mapping_file:
            self.mapping = yaml.safe_load(
                mapping_file.read()
            )
        self.ignored = self.mapping.get('ignored', [])
        print('Mapování nahráno')

    def needs_refresh(self, hltb_query_ts: str) -> bool:
        '''Whether a game with particular timestamp needs to be queried again'''
        if not hltb_query_ts:
            return True
        if self.args.cache_ttl <= 0:
            return True
        cache_ts = datetime.datetime.fromisoformat(hltb_query_ts)
        now_ts = datetime.datetime.now(datetime.timezone.utc)
        ttl = datetime.timedelta(days=self.args.cache_ttl)
        return (now_ts - cache_ts) > ttl

    @staticmethod
    def bool2str(value: bool) -> str:
        if value is True:
            return 'ano'
        elif value is False:
            return 'ne'
        else:
            return str(value)

    @classmethod
    def equal_names(cls, names1: Iterable[str], names2: Iterable[str]) -> bool:
        '''Return whether two sets of game names are basically the same, just differing in
        unimportant details like letter casing or an extra colon. Returns True if any name from
        the first set is equal to any name from the second set. (Using sets allows you to specify
        e.g. game name aliases).'''
        def equalize(nameset: Iterable[str]) -> Generator[str, None, None]:
            for name in nameset:
                name = cls.more_searchable_name(name)
                name = name.casefold()
                yield name

        names1 = list(equalize(names1))
        names2 = list(equalize(names2))

        for name1 in names1:
            for name2 in names2:
                if name1 == name2:
                    return True
        return False

    @staticmethod
    def more_searchable_name(name: str) -> str:
        '''Convert a game name to a more searchable variant, i.e. remove or
        replace characters which HLTB has a problem with (like an extra colon).
        '''
        name = name.replace(':', ' ')
        name = name.replace('–', ' ')
        name = name.replace('-', ' ')
        name = ' '.join(name.split())  # remove multiple spaces
        name = name.strip()
        return name

    def export_table_data(self) -> List[List[str]]:
        link_template = 'https://www.databaze-her.cz/h{}'
        header = ['Název', 'Rok', 'HLTB Main', 'HLTB Extra', 'HLTB Complete', 'HLTB All Styles',
                  'Chci hrát', 'Dohráno', 'Vlastněno', 'Odkaz']

        table = []
        # header
        table.append(header)
        # data fields
        for game in self.games:
            if game.dh_id in self.ignored and not self.args.include_ignored:
                continue
            link = link_template.format(game.dh_id)
            finished = game.finished_ts or self.bool2str(game.finished)
            table.append(
                [game.title, game.year, game.time_main or '', game.time_extra or '',
                 game.time_complete or '', game.time_all or '', self.bool2str(game.wantplay),
                 finished, self.bool2str(game.owned), link]
            )

        return table

    def export(self) -> None:
        if self.args.output.endswith('.xlsx'):
            self.export_xlsx()
        elif self.args.output.endswith('.ods'):
            self.export_ods()
        elif self.args.output.endswith('.csv'):
            self.export_csv()
        else:
            raise RuntimeError("Nepodporovaná přípona souboru (nemělo by se stát): "
                f'{self.args.output}')

        colorprint(Color.SUCCESS, 'Výsledný soubor uložen: {} ✔'.format(
            os.path.abspath(self.args.output)))

    def export_csv(self) -> None:
        print('Ukládám výsledky do CSV ...')
        table = self.export_table_data()
        with open(self.args.output, 'w', newline='') as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerows(table)

    def export_ods(self) -> None:
        print('Ukládám výsledky do ODS ...')
        # save main values
        table = self.export_table_data()
        sheet = pyexcel.Sheet(name='stats', sheet=table, name_columns_by_row=0)
        sheet.save_as(self.args.output)
        # add formulas (pyexcel can't do those, use ezodf instead)
        doc = ezodf.opendoc(self.args.output)
        doc.backup = False
        sheet = doc.sheets[0]
        sheet.append_rows(2)
        lastrow = sheet.nrows() - 1
        sheet[lastrow, 2].formula = '=SUBTOTAL(9;C:C)'
        sheet[lastrow, 3].formula = '=SUBTOTAL(9;D:D)'
        sheet[lastrow, 4].formula = '=SUBTOTAL(9;E:E)'
        sheet[lastrow, 5].formula = '=SUBTOTAL(9;F:F)'
        doc.save()

    def export_xlsx(self) -> None:
        print('Ukládám výsledky do XLSX ...')
        table = self.export_table_data()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'stats'
        for row in table:
            sheet.append(row)
        # add hyperlinks to 'Link' column
        for sheet_slice in sheet['J2:J{}'.format(sheet.max_row)]:
            cell = sheet_slice[0]
            cell.hyperlink = cell.value
        # add auto filter
        sheet.auto_filter.ref = sheet.dimensions
        # add Sums
        sheet.append([])
        sheet.append(['Celkem (hodiny):', '', '=SUBTOTAL(109,C:C)', '=SUBTOTAL(109,D:D)',
                     '=SUBTOTAL(109,E:E)', '=SUBTOTAL(109,F:F)'])
        sheet.cell(sheet.max_row, 1).font = openpyxl.styles.Font(bold=True)
        # set styles
        sheet.column_dimensions['C'].number_format = \
            sheet.column_dimensions['D'].number_format = \
            sheet.column_dimensions['E'].number_format = '# ##0'
        sheet.row_dimensions[1].font = openpyxl.styles.Font(bold=True)
        sheet.row_dimensions[1].alignment = openpyxl.styles.Alignment(
            horizontal='center')
        sheet.column_dimensions['J'].font = openpyxl.styles.Font(
            underline='single')
        # freeze views
        sheet.freeze_panes = 'A2'
        # resize columns
        colwidths = [('A', 45), ('B', 8), ('C', 14), ('D', 14), ('E', 18),
                     ('F', 18), ('G', 13), ('H', 12), ('I', 13), ('J', 35)]
        for colname, width in colwidths:
            sheet.column_dimensions[colname].width = width

        workbook.save(self.args.output)


class HLTBError(BaseException):
    pass


class Color(enum.Enum):
    ERROR = colorama.Fore.RED + colorama.Style.BRIGHT
    ERROR_DETAILS = colorama.Fore.RED
    INFO = colorama.Fore.BLUE + colorama.Style.BRIGHT
    SUCCESS = colorama.Fore.GREEN + colorama.Style.BRIGHT


def colorprint(color: Color, msg: str, **kwargs):
    print(f'{color.value}{msg}', **kwargs)

def print_error(msg: str):
    colorprint(Color.ERROR, msg, file=sys.stderr)

def find_prog_dir() -> str:
    return os.path.dirname(os.path.realpath(__file__))


if __name__ == '__main__':
    args = parse_args()
    colorama.init(autoreset=True)
    games = create_dh_gamelist(html_filenames=args.html_filenames)
    hltb = HLTB(args=args, games=games)
    try:
        hltb.run()
    except KeyboardInterrupt:
        print_error('SIGINT zachycen, ukládám dočasnou paměť a končím!')
        hltb.save_cache()
        sys.exit(1)
    except HLTBError:
        print_error('Detekována chyba spojení, prosím zkuste to znovu!')
        hltb.save_cache()
        sys.exit(1)
    except BaseException as e:
        hltb.save_cache()
        raise e
    hltb.export()
